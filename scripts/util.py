import openpyxl as xl
from openpyxl.styles import Font, Alignment


def fill(ws, name):

    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20


    ws.merge_cells(range_string='A1:E1')

    cell = ws['A1']
    cell.value = f"{name}"
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(size=23, color='000000', bold=True)


    cell = ws['A3']
    cell.value = 'Name'
    cell.font = Font(size=11, color='000000', bold=True)

    cell = ws['B3']
    cell.value = 'Class'
    cell.font = Font(size=11, color='000000', bold=True)

    cell = ws['C3']
    cell.value = 'Section'
    cell.font = Font(size=11, color='000000', bold=True)

    cell = ws['D3']
    cell.value = 'Father Name'
    cell.font = Font(size=11, color='000000', bold=True)

    cell = ws['E3']
    cell.value = 'Village'
    cell.font = Font(size=11, color='000000', bold=True)

    cell = ws['F3']
    cell.value = 'Phone Number'
    cell.font = Font(size=11, color='000000', bold=True)

    cell = ws['G3']
    cell.value = 'Photo '
    cell.font = Font(size=11, color='000000', bold=True)


    return ws
