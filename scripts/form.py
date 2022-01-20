from tkinter import Tk, Label, Button, messagebox, filedialog
from tkinter import *
from tkinter import ttk
import sys, os, glob
import openpyxl as xl
from util import fill
import cv2


flag_form = 0

def visible(win):
    global flag_form

    if flag_form == 0:
        flag_form = 1
        win.withdraw()
    else:
        flag_form = 0
        win.deiconify()


def click_photo(name):
    
    print("clicking photo at", os.getcwd())
    os.chdir('photos')
    print(name)
    webcam = cv2.VideoCapture(0)
    
    while True:

        ret,photo = webcam.read()
        cv2.imshow('Frame', photo)

        key = cv2.waitKey(1)

        if key == ord('s'):
            file = f"{name}.jpg"
            webcam.release()
            cv2.destroyAllWindows()
            cv2.imwrite(filename=file, img=photo)
            flags = 1
            break
            
        elif key == ord('q'):
            webcam.release()
            cv2.destroyAllWindows()
            break

    if flags == 1:
        messagebox.showinfo("Message", "Photo clicked successfully!")
    else:
        messagebox.showinfo("Message", "Unsuccesful attempt!")
            
    return
            

def populate(win, name, num, fname, cls, sec, pid, vil):
    print(os.getcwd())
    name = name.get()
    num = num.get()
    fname = fname.get()
    cls = cls.get()
    sec = sec.get()
    pid = pid.get()
    vil = vil.get()

    filename = os.path.basename(os.getcwd())
    
    
    files = os.listdir()
    print("Checking")
    for file in files:
        if '.xlsx' in file:
            wb = xl.load_workbook(file)
    
            

    sheet_names = wb.sheetnames

    for sheet in sheet_names:
        if sheet == cls:
            ws = wb[sheet]
    else:
        ws = wb.create_sheet(cls)
        ws = fill(ws, filename)

    row = ws.max_row + 1

    cell = ws[f'A{row}']
    cell.value = name

    cell = ws[f'B{row}']
    cell.value = cls

    cell = ws[f'C{row}']
    cell.value = sec

    cell = ws[f'D{row}']
    cell.value = fname

    cell = ws[f'E{row}']
    cell.value = vil

    cell = ws[f'F{row}']
    cell.value = num

    cell = ws[f'G{row}']
    cell.value = pid

    wb.save(f"{filename}.xlsx")
    print("File saved at", os.getcwd())

    win.quit()

    click_photo(pid)

    
    


class Form:

    def __init__(self, main_win, win=Tk()):

        self.win = Toplevel(main_win)
        win.protocol("WM_DELETE_WINDOW",self.event_X)
        win.title("Enter details - Photosoft 3.0")
        win.geometry("450x450")
        #win.geometry("450x230+450+170")

     #variables
        name = StringVar(win)
        num = StringVar(win)
        fname = StringVar(win)
        cls = StringVar(win)
        sec = StringVar(win)
        pid = StringVar(win)
        vil = StringVar(win)

    #labels
    
        self.nameLabel = Label(win,text="Name:")
        self.nameLabel.place(relx=0.285, rely=0.098, height=20, width=70)

        self.classLabel = Label(win, text="Class:")
        self.classLabel.place(relx=0.285, rely=0.208, height=20, width=70)

        self.secLabel = Label(win, text="Section:")
        self.secLabel.place(relx=0.285, rely=0.318, height=20, width=70)

        self.fnameLabel = Label(win,text="Father's name:")
        self.fnameLabel.place(relx=0.220, rely=0.428, height=20, width=100)

        self.vLabel = Label(win,text="Village :")
        self.vLabel.place(relx=0.220, rely=0.538, height=20, width=100)

        self.numLabel = Label(win,text="Phone Number :")
        self.numLabel.place(relx = 0.220, rely =0.648, height=20, width =100)

        self.idLabel = Label(win,text="Photo Id: ")
        self.idLabel.place(relx = 0.220, rely = 0.758, height=20, width =100)

    


    #entry

        self.name_box = ttk.Entry(win, textvariable=name)
        self.name_box.place(relx=0.440, rely=0.098, height=20, relwidth=0.35)

        self.class_box = ttk.Entry(win, textvariable=cls)
        self.class_box.place(relx=0.440,rely=0.208 , height=20, relwidth=0.35)

        self.sec_box = ttk.Entry(win, textvariable=sec)
        self.sec_box.place(relx=0.440,rely=0.318 , height=20, relwidth=0.35)

        self.fname_box = ttk.Entry(win, textvariable=fname)
        self.fname_box.place(relx=0.440, rely=0.428, height=20, relwidth=0.35)

        self.v_box = ttk.Entry(win, textvariable=vil)
        self.v_box.place(relx=0.440, rely=0.538, height=20, relwidth=0.35)

        self.num_box = ttk.Entry(win, textvariable=num)
        self.num_box.place(relx=0.440, rely=0.648, height=20, relwidth=0.35)

        self.id_box = ttk.Entry(win, textvariable=pid)
        self.id_box.place(relx=0.440, rely=0.758, height=20, relwidth=0.35)

        self.b = ttk.Button(win,text="submit", command=lambda: populate(win, name, num, fname, cls, sec, pid, vil))
        self.b.place(relx=0.400, rely=0.868)

    def event_X(self):
            sys.exit()

    def mainloop_window(self):
            self.win.mainloop()
   


